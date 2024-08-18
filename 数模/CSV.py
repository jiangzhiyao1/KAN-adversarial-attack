import openpyxl
import csv

def excel_to_csv_large_file(input_excel_file, output_csv_file, sheet_name=0, chunk_size=100000):
    # 打开Excel文件
    wb = openpyxl.load_workbook(input_excel_file, read_only=True)
    
    # 如果sheet_name是整数，则获取对应的工作表名称
    if isinstance(sheet_name, int):
        sheet_names = wb.sheetnames
        if sheet_name < len(sheet_names):
            ws = wb[sheet_names[sheet_name]]
        else:
            raise IndexError(f"Worksheet index {sheet_name} is out of range.")
    else:
        ws = wb[sheet_name]

    # 打开CSV文件以写入模式
    with open(output_csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        
        # 获取Excel文件的表头
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        writer.writerow(header)

        # 逐块写入数据
        data_chunk = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
            data_chunk.append(row)
            if i % chunk_size == 0:
                writer.writerows(data_chunk)
                data_chunk = []

        # 写入剩余数据
        if data_chunk:
            writer.writerows(data_chunk)

# 调用函数，指定输入的Excel文件路径和输出的CSV文件路径
input_excel_file = '附件4.xlsx'
output_csv_file = '附件4.csv'

excel_to_csv_large_file(input_excel_file, output_csv_file)
